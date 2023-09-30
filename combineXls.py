from openpyxl import Workbook, load_workbook
from HajjPdfToXl import stations, alramyat_colors, cell_style, adjust_col_width
import os


def create_master_headings(master_path):
    table_headings = ["اسم المراقب","اسم المشرف","اسم المعاون",
                  'عدد الحجاج',"اوقات التفويج","يوم التفويج",
                  "الفوج","#","الرمية","رقم المركز"]
    wb_master = Workbook()
    ws_master = wb_master.active
    ws_master.append(table_headings)
    adjust_col_width(ws_master)
    cell_style(ws_master, "D9D9D9", 1, "DIN Next LT Arabic Regular", 12, True, "EC733A")
    wb_master.save(master_path)

xl_path = "Result Excel//"
ws_xl_name = "Sheet"
master_path = "master.xlsx"
ws_master_name = "Sheet"
master_row = 1

if(os.path.isfile(master_path)):
    os.remove(master_path)
create_master_headings(master_path)

for station_number in stations:
    wb_master = load_workbook(master_path)
    wb_xl = load_workbook(xl_path+str(station_number)+".xlsx")
    ws_master = wb_master.active 

    ws_master = wb_master.active
    ws_xl = wb_xl.active


    for row in range(2, ws_xl.max_row+1):        
        master_row+=1
        for col in range(1, ws_xl.max_column+1):
            ws_master.cell(row=master_row, column=col).value = ws_xl.cell(row=row, column=col).value
        alramyah = int((ws_master.cell(row=master_row, column=9).value)[-1]) -1
        cell_style(ws_master, alramyat_colors[alramyah], master_row)    
    
    wb_master.save(master_path)
    wb_xl.save(xl_path+str(station_number)+".xlsx")

