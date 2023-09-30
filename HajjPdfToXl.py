from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
import tabula
import csv
import os

# Functions Area
# csv الى pdf استخراج وتحويل ملف 
# تحويل من ص 2 الى ص 5
def  ExtractPdfFile(pdf_path, pdf_name, csv_folder_container):
    tabula.convert_into(pdf_path, f"{csv_folder_container}{pdf_name}.csv", 
                        output_format="csv", pages="all", stream=True)
    
# تعديل خلايا الجدول
def cell_style(ws, bg_color, current_row, 
               font_name="Arial", font_size=12, font_bold=False, font_color="212529"):
    
    for rows in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type = "solid")
            cell.font = Font(name=font_name, size=font_size, bold=font_bold, color=font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(border_style="thin",
                           color='212529'),
                 right=Side(border_style="thin",
                            color='212529'),
                 top=Side(border_style="thin",
                          color='212529'),
                 bottom=Side(border_style="thin",
                             color='212529'),)

# تعديل عرض الاعمدة
def adjust_col_width(ws):
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 23
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10

def init():
    # اسماء المعاونين
    sultan_dmnhori = [1,2,3,4,5,6,9,"777B"] 
    associate_dict = dict.fromkeys(sultan_dmnhori, "سلطان دمنهوري")

    # اسماء المشرفين
    mansour_bahmdan = [1,2]
    anmar = [3,4,5]
    morad = [6,9, "777B"]
    supervisor_dict = dict.fromkeys(mansour_bahmdan, "منصور باحمدان")
    supervisor_dict.update(dict.fromkeys(anmar, "انمار"))
    supervisor_dict.update(dict.fromkeys(morad, "مراد"))

    # اسماء المراقبين
    observer_dict = {}

    # البيانات الاضافيةالمدخلة يدويا 
    alramyat = ["الرمية الأولى - 1","الرمية الثانية - 2","الرمية الثالثة - 3","الرمية الرابعة - 4"]
    # [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20, 
    # 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 
    # 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 
    # 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 
    # 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 
    # 101, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 501,
    # 502, 503, 504, 701, 702, 703, 704, 705, 706, 707, 777]

    table_headings = ["اسم المراقب","اسم المشرف","اسم المعاون",
                    'عدد الحجاج',"اوقات التفويج","يوم التفويج",
                    "الفوج","#","الرمية","رقم المركز"]

    # csv مجلد ملف ال 
    csv_path = "Result CSV//"
    xl_path = "Result Excel//"

    #بداية الكود
    for station_number in stations:
        pdf_path = "Data//"+str(station_number)+".pdf"
        pdf_name = str(station_number)

        if(os.path.isfile(csv_path+pdf_name+".csv")):
            print(pdf_name+".csv Exist")
        else:
            ExtractPdfFile(pdf_path, pdf_name, csv_path)

        if(os.path.isfile(xl_path+pdf_name+".xlsx")):
            print(pdf_name+".xlsx Exist")
            continue


        wb = Workbook()
        ws = wb.active
        
        with open(f'{csv_path}{pdf_name}.csv') as f:
            reader = csv.reader(f, delimiter=',')
            header_added = False
            row_count = 1
            # التأكد من يوم الرمية (لان بعض المكاتب ليس لها رمية اولى)
            next(reader)
            if(next(reader)[2] == str(11)):
                alramyh = 0
            else:
                alramyh = -1
            f.seek(0)
            for row in reader:
                # مسح رؤوس الاعمدة المكررة
                if("?" in row[0]):
                    alramyh+=1
                    # اضافة رؤوس الاعمدة
                    if(header_added != True):
                        row = table_headings
                        header_added = True
                        ws.append(row)
                        cell_style(ws, "D9D9D9", row_count, "DIN Next LT Arabic Regular", 12, True, "EC733A")
                        adjust_col_width(ws)
                                
                    continue    
                # اضافة المعلومات المدخلة يدويا
                row_count+=1
                row.insert(0, "")
                row.insert(1, supervisor_dict[station_number] if supervisor_dict[station_number] is not None  else "")
                row.insert(2, associate_dict[station_number] if associate_dict[station_number] is not None else "")
                row.append(alramyat[alramyh])
                row.append(station_number)  
                ws.append(row)
                cell_style(ws, alramyat_colors[alramyh], row_count)
        
            tab = Table(displayName="Table1", ref="A1:J"+str(row_count))
            ws.add_table(tab)
        wb.save(f'{xl_path}{pdf_name}.xlsx')

stations = [1,2,3,4,5,6,9,"777B"]
alramyat_colors = ["C6E0B4","BDD7EE","FFE699","F8CBAD"]
    
if(__name__ == "__main__"):
    init()